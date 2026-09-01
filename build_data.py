#!/usr/bin/env python3
"""Parse the National Bible Bee passage test-bank markdown files into the game's
question data (assets/data.js -> window.QUIZ_DATA).

Usage:  python3 build_data.py
"""
import re, json, os

BANKS = {
    "primary": "Primary_Passages_Test_Bank.md",
    "junior":  "Junior_Passages_Test_Bank.md",
    "senior":  "Senior_Passages_Test_Bank.md",
}

def clean(s):
    # normalise the non-breaking hyphen used in answer keys
    return s.replace("‑", "-").strip()

def derive_category(title):
    """Map a set's heading to a (category-key, study-chapter) pair."""
    t = title.lower()
    if "random access" in t:
        chap = None
        if "random access," in t:
            chap = title.split("Random Access,", 1)[1].strip()
            if "mixed" in chap.lower():
                chap = "Mixed"
        return "random", chap
    if "cross-reference" in t or "cross reference" in t: return "xref", None
    if "feast" in t: return "feasts", None
    if "unique words" in t: return "unique", None
    if "exactly two" in t: return "words2", None
    if "three or four" in t or "exactly three" in t: return "words3", None
    if "geographical" in t: return "geo", None
    if "names for god" in t or "titles of the persons" in t or "trinity" in t: return "names", None
    if "ten commandments" in t: return "commandments", None
    if "theme" in t: return "theme", None
    return "other", None

def parse(path):
    lines = open(path, encoding="utf-8").read().splitlines()
    sets, keys, cur, i = {}, {}, None, 0
    titles = {}
    qmark   = re.compile(r"^\*\*(\d+)\.\*\*\s*(.*)$")
    setmark = re.compile(r"^#\s*SET\s+(\d+)\s*(?:[—-]\s*(.*))?$", re.I)
    keymark = re.compile(r"Answer key\s*[—-]\s*SET\s+(\d+)\s*:\s*(.*)$", re.I)
    while i < len(lines):
        line = lines[i]
        ms = setmark.match(line)
        if ms:
            cur = int(ms.group(1)); sets.setdefault(cur, [])
            titles[cur] = (ms.group(2) or "").strip()
            i += 1; continue
        mk = keymark.search(line)
        if mk:
            s = int(mk.group(1)); d = {}
            for pair in clean(mk.group(2)).split(","):
                m2 = re.search(r"(\d+)-([A-D])", clean(pair))   # search, not match:
                if m2:                                          # first pair is bolded
                    d[int(m2.group(1))] = m2.group(2)
            keys[s] = d; i += 1; continue
        mq = qmark.match(line)
        if mq and cur is not None:
            qnum = int(mq.group(1)); qtext = mq.group(2).strip()
            j = i + 1
            while j < len(lines) and lines[j].strip() == "":
                j += 1
            optline = lines[j] if j < len(lines) else ""
            parts = re.split(r"\s*(?:^|\s)([A-D])\)\s*", optline)
            opts = {}
            for k in range(1, len(parts) - 1, 2):
                opts[parts[k]] = parts[k + 1].strip()
            sets[cur].append({"num": qnum, "q": qtext, "opts": opts})
            i = j + 1; continue
        i += 1
    out = []
    for s, ql in sets.items():
        kd = keys.get(s, {})
        for q in ql:
            L = kd.get(q["num"])
            if not L or set(q["opts"].keys()) != {"A", "B", "C", "D"}:
                continue
            order = ["A", "B", "C", "D"]
            out.append({"q": q["q"], "o": [q["opts"][x] for x in order],
                        "a": order.index(L), "s": s})
    set_meta = {}
    for s, title in titles.items():
        cat, chap = derive_category(title)
        set_meta[str(s)] = {"t": title, "c": cat, "ch": chap}
    return out, (max(sets) if sets else 0), set_meta

# ============================================================
#  XLSX test banks (Distinctive Words / Direct Quotes workbooks)
# ============================================================
XLSX_DIR = "xlsx_banks"
DIVS = ("primary", "junior", "senior")

# Each subsheet type is its own game category (forward + reverse share a kind, so
# they merge automatically). Greek (Matthew) and Hebrew (Proverbs) word-study types
# are kept distinct so the names can say which language.
def sheet_kind(sheet):
    s = sheet.lower()
    if "nkjv-greek" in s: return "gkword"
    if "nkjv-hebrew" in s: return "hbword"
    if "verse location" in s: return "verseloc"
    if "strongs-greek" in s: return "gkstrong"
    if "strongs-hebrew" in s: return "hbstrong"
    if "vines" in s: return "gkvine"        # Vine's definitions (Greek / Matthew)
    if "outline" in s: return "hboutline"   # Outline definitions (Hebrew / Proverbs)
    if "section refs" in s: return "sectionref"
    if "which is not a cr" in s or re.search(r"\bref \(", s): return "wordxref"
    if "section titles" in s: return "sectitle"
    if "decalogue" in s: return "decalogue"
    if "ot legal" in s: return "otlegal"
    if "no-parallel" in s: return "noparallel"
    if "crossref to passage" in s or "passage to crossref" in s: return "provxref"
    if ("crossref to matt" in s or "matt passage to crossref" in s
            or "to matt" in s or "matt to" in s): return "parallel"
    return None

def sheet_division(sheet):
    s = sheet.lower()
    if "junior" in s: return "junior"
    if "primary" in s: return "primary"
    if "senior" in s: return "senior"
    return None

def file_chapter(fname):
    m = re.match(r"(Matthew|Proverbs)(\d)", os.path.basename(fname))
    return f"{m.group(1)} {m.group(2)}" if m else None

def text_chapter(text):
    t = str(text)
    m = re.search(r"Matt(?:hew|\.)?\s*([567])\b", t)
    if m: return "Matthew " + m.group(1)
    m = re.search(r"Prov(?:erbs|\.)?\s*([34])\b", t)
    if m: return "Proverbs " + m.group(1)
    return None

def parse_xlsx(here):
    """Return {division: [question dicts]} from the xlsx workbooks, or {} if the
    folder or the openpyxl dependency is absent."""
    xdir = os.path.join(here, XLSX_DIR)
    if not os.path.isdir(xdir):
        return {}
    try:
        import openpyxl
    except ImportError:
        print("  ! xlsx banks present but openpyxl is not installed — skipping them.")
        print("    Run:  pip install openpyxl   (then re-run build_data.py)")
        return {}
    import glob
    out = {d: [] for d in DIVS}
    for f in sorted(glob.glob(os.path.join(xdir, "*.xlsx"))):
        filech = file_chapter(f)
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        for sh in wb.sheetnames:
            if sh == "Overview":
                continue
            kind = sheet_kind(sh)
            if not kind:
                continue
            cat = kind                     # one category per subsheet type
            sdiv = sheet_division(sh)
            for row in wb[sh].iter_rows(min_row=2, values_only=True):
                if not row or len(row) < 6:
                    continue
                q, a, b, c, d, letter = row[0], row[1], row[2], row[3], row[4], row[5]
                opts = [a, b, c, d]
                if not q or letter not in ("A", "B", "C", "D"):
                    continue
                if any(o is None or str(o).strip() == "" for o in opts):
                    continue
                ans = row[6] if len(row) > 6 else ""
                ch = filech or text_chapter(str(ans) + " " + str(q))
                divs = [sdiv] if sdiv else list(DIVS)
                if ch == "Proverbs 4":                    # Primary doesn't study Prov 4
                    divs = [x for x in divs if x != "primary"]
                item = {"q": str(q).strip(), "o": [str(o).strip() for o in opts],
                        "a": ["A", "B", "C", "D"].index(letter), "c": cat}
                if ch:
                    item["ch"] = ch
                for dv in divs:
                    out[dv].append(dict(item))
        wb.close()
    return out

def main():
    here = os.path.dirname(os.path.abspath(__file__))
    data, meta = {}, {}
    for div, fname in BANKS.items():
        qs, maxset, set_meta = parse(os.path.join(here, fname))
        data[div] = qs
        meta[div] = {"sets": maxset, "count": len(qs), "setMeta": set_meta}
        print(f"{div:8s} {len(qs):5d} md questions across {maxset} sets")

    # merge in the xlsx workbooks (new categories, tagged with c/ch directly)
    xlsx = parse_xlsx(here)
    for div in BANKS:
        add = xlsx.get(div, [])
        if add:
            data[div].extend(add)
            meta[div]["count"] = len(data[div])
            print(f"{div:8s} +{len(add):5d} xlsx questions  ->  {len(data[div])} total")
    payload = {"meta": meta, "data": data}
    js = "window.QUIZ_DATA=" + json.dumps(payload, ensure_ascii=False,
                                          separators=(",", ":")) + ";"
    out_path = os.path.join(here, "assets", "data.js")
    open(out_path, "w", encoding="utf-8").write(js)
    print(f"wrote {out_path} ({len(js):,} bytes, "
          f"{sum(len(v) for v in data.values())} questions total)")

if __name__ == "__main__":
    main()
